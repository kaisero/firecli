import copy
import json
from datetime import datetime
from logging import getLogger
from typing import Dict, List

from benedict import benedict
from fireREST import FMC
from fireREST.mapping import ICMP_TYPE, IP_PROTOCOL, STATE
from fireREST.exceptions import GenericApiError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from firecli.api.afa import AFA
from firecli.api.compliance import ZoneCompliance

logger = getLogger(__name__)


class API(object):
    def __init__(self, cfg: Dict):
        self.cfg = cfg
        self._afa = None
        self._fmc = None

    @property
    def fmc(self):
        if not self._fmc:
            self._fmc = FMC(
                hostname=self.cfg['fmc']['hostname'],
                username=self.cfg['fmc']['username'],
                password=self.cfg['fmc']['password'],
                domain=self.cfg['fmc']['domain'],
                timeout=self.cfg['fmc']['timeout'],
                dry_run=self.cfg['dry_run'],
            )
        return self._fmc

    @property
    def afa(self):
        if not self._afa:
            self._afa = AFA(
                hostname=self.cfg['afa']['hostname'],
                username=self.cfg['afa']['username'],
                password=self.cfg['afa']['password'],
                timeout=self.cfg['afa']['timeout'],
            )
        return self._afa

    @staticmethod
    def filtered_accessrules(policy_id: str, accessrules: List, policy_filter: str):
        if policy_filter == 'parent':
            return [
                accessrule for accessrule in accessrules if accessrule['metadata']['accessPolicy']['id'] != policy_id
            ]
        if policy_filter == 'child':
            return [
                accessrule for accessrule in accessrules if accessrule['metadata']['accessPolicy']['id'] == policy_id
            ]
        return accessrules

    @staticmethod
    def csv_squashed(csv: List):
        empty_fields = [True for name in csv[0].split(';')]
        for row_index, row in enumerate(csv):
            if row_index > 0:
                fields = row.split(';')
                for field_index, field in enumerate(fields):
                    if field != '':
                        empty_fields[field_index] = False
        for row_index, row in enumerate(csv):
            fields = [field for field_index, field in enumerate(row.split(';')) if not empty_fields[field_index]]
            csv[row_index] = ';'.join(fields)
        return csv

    @staticmethod
    def csv_header(obj_type: str):
        fields = {
            'accessrule': [
                'Section',
                'Category',
                'Name',
                'Source Zones',
                'Dest Zones',
                'Source Networks',
                'Dest Networks',
                'VLAN Tags',
                'Users',
                'Applications',
                'Source Ports',
                'Dest Ports',
                'URLs',
                'Source SGT',
                'Dest SGT',
                'Action',
                'Variable Set',
                'Intrusion Policy',
                'File Policy',
                'SafeSearch',
                'Logging',
                'Comments',
                'Hitcount',
                'First Hitcount',
                'Last Hitcount',
            ]
        }
        return ';'.join(fields[obj_type])

    def accessrule_to_csv(self, data: Dict):
        section = f'{data["metadata"]["section"]} - {data["metadata"]["accessPolicy"]["name"]}'
        category = data['metadata']['category']
        name = data['name']
        src_zones = self.zone_to_csv(data=data['sourceZones']) if 'sourceZones' in data else ''
        dst_zones = self.zone_to_csv(data=data['destinationZones']) if 'destinationZones' in data else ''
        src_networks = self.network_to_csv(data=data['sourceNetworks']) if 'sourceNetworks' in data else ''
        dst_networks = self.network_to_csv(data=data['destinationNetworks']) if 'destinationNetworks' in data else ''
        vlan_tags = self.vlan_to_csv(data=data['vlanTags']) if 'vlanTags' in data else ''
        users = self.user_to_csv(data=data['users']) if 'users' in data else ''
        applications = self.app_to_csv(data=data['applications']) if 'applications' in data else ''
        src_ports = self.port_to_csv(data=data['sourcePorts']) if 'sourcePorts' in data else ''
        dst_ports = self.port_to_csv(data=data['destinationPorts']) if 'destinationPorts' in data else ''
        urls = self.url_to_csv(data=data['urls']) if 'urls' in data else ''
        src_sgt = ''
        dst_sgt = ''
        action = data.get('action', '')
        variable_set = data['variableSet']['name'] if 'variableSet' in data else ''
        ips_policy = data['ipsPolicy']['name'] if 'ipsPolicy' in data else ''
        file_policy = data['filePolicy']['name'] if 'filePolicy' in data else ''
        safesearch = STATE[data['safeSearch']['enabled']] if 'safeSearch' in data else ''
        logging = self.logging_to_csv(data)
        comments = self.comment_to_csv(data=data['commentHistoryList']) if 'commentHistoryList' in data else ''
        hitcount = data['hitcount']['hitCount'] if 'hitcount' in data else '0'
        hitcount_first_match = data['hitcount']['firstHitTimeStamp'] if 'hitcount' in data else ''
        hitcount_last_match = data['hitcount']['lastHitTimeStamp'] if 'hitcount' in data else ''
        fields = [
            section,
            category,
            name,
            src_zones,
            dst_zones,
            src_networks,
            dst_networks,
            vlan_tags,
            users,
            applications,
            src_ports,
            dst_ports,
            urls,
            src_sgt,
            dst_sgt,
            action,
            variable_set,
            ips_policy,
            file_policy,
            safesearch,
            logging,
            comments,
            hitcount,
            hitcount_first_match,
            hitcount_last_match,
        ]
        return ';'.join([str(item) for item in fields])

    @staticmethod
    def app_to_csv(data):
        apps = list()
        if 'applications' in data:
            for item in data['applications']:
                apps.append(f'{item["name"]} ({item["id"]})')
        if 'applicationFilters' in data:
            for item in data['applicationFilters']:
                apps.append(f'{item["name"]} (ApplicationFilter)')
        if 'inlineApplicationFilters' in data:
            for _item in data['inlineApplicationFilters']:
                apps.append('Unparseable Filter')
        return ','.join(apps)

    def network_to_csv(self, data):
        result = list()
        if 'literals' in data:
            for item in data['literals']:
                result.append(item['value'])
        if 'objects' in data:
            for item in data['objects']:
                result.append(self.to_csv_item_recursive(item))
        return ','.join(result)

    @staticmethod
    def comment_to_csv(data):
        result = list()
        for item in data:
            result.append(f'{item["user"]["name"]} on {item["date"]}: {item["comment"]}')
        return ','.join(result)

    @staticmethod
    def logging_to_csv(data):
        result = list()
        if data['sendEventsToFMC']:
            result.append('FMC')
        if data['enableSyslog']:
            result.append('Syslog')
        return ','.join(result)

    def port_to_csv(self, data):
        result = list()
        if 'literals' in data:
            for item in data['literals']:
                if 'icmpType' in item:
                    icmp_type = 'ICMP/Any'
                    try:
                        icmp_type = f'ICMP/{ICMP_TYPE[int(item["icmpType"])]}'
                    except ValueError:
                        pass
                    result.append(icmp_type)
                else:
                    port = f'/{item.get("port", "")}'
                    result.append(f'{IP_PROTOCOL[item["protocol"]]}{port}')
        if 'objects' in data:
            for item in data['objects']:
                result.append(self.to_csv_item_recursive(item))
        return ','.join(result)

    def url_to_csv(self, data):
        result = list()
        if 'literals' in data:
            for item in data['literals']:
                result.append(item['url'])
        if 'objects' in data:
            for item in data['objects']:
                result.append(self.to_csv_item_recursive(item))
        return ','.join(result)

    @staticmethod
    def user_to_csv(data):
        result = list()
        if 'literals' in data:
            for item in data['literals']:
                result.append(item['value'])
        if 'objects' in data:
            for item in data['objects']:
                result.append(item['name'])
        return ','.join(result)

    @staticmethod
    def vlan_to_csv(data):
        result = list()
        if 'literals' in data:
            for item in data['literals']:
                vlan_id = (
                    f'{item["startTag"]}-{item["endTag"]}'
                    if item['startTag'] != item['endTag']
                    else str(item['startTag'])
                )
                result.append(vlan_id)
        if 'objects' in data:
            for item in data['objects']:
                result.append(item['name'])
        return ','.join(result)

    @staticmethod
    def zone_to_csv(data):
        result = list()
        if 'literals' in data:
            for item in data['literals']:
                result.append(item['value'])
        if 'objects' in data:
            for item in data['objects']:
                result.append(item['name'])
        return ','.join(result)

    def to_csv_item_recursive(self, data: Dict):
        if data['type'] not in ('NetworkGroup', 'PortObjectGroup'):
            if 'value' in data:
                return f'{data["name"]} ({data["value"]})'
            if 'protocol' in data:
                port = f'/{data.get("port", "")}'
                return f'{data["name"]} ({data["protocol"]}{port})'
            if 'icmpType' in data:
                if data['icmpType'].isnumeric():
                    return f'{data["name"]} (ICMP/{ICMP_TYPE[int(data["icmpType"])]})'
                else:
                    return f'{data["name"]} (ICMP)'
            if 'url' in data:
                return f'{data["name"]} ({data["url"]})'
            return data['name']
        items = list()
        if 'objects' in data:
            for item in data['objects']:
                obj = f'{data["name"]} ({self.to_csv_item_recursive(item)})'
                items.append(obj)
        if 'literals' in data:
            for item in data['literals']:
                lit = f'{data["name"]} ({item["value"]})'
                items.append(lit)
        return ','.join(items)

    def to_excel(self, obj_type: str, data):
        workbook = Workbook()
        if obj_type == 'accesspolicy':
            worksheet = workbook.active
            for row_id, row in enumerate(data, start=1):
                for col_id, col_value in enumerate(row.split(';'), start=1):
                    cell = worksheet.cell(row=row_id, column=col_id)
                    cell.value = col_value
                    self.style_cell(cell)
            worksheet.auto_filter.ref = worksheet.dimensions
            self.freeze_excel_toprow(worksheet)
            self.fix_column_dimensions(worksheet)
        return workbook

    @staticmethod
    def style_cell(cell):
        if (cell.row % 2) != 0:
            cell.fill = PatternFill(start_color='004B8C', end_color='004B8C', fill_type='solid')
            if cell.row != 1:
                cell.font = Font(
                    name='Consolas',
                    size=11,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FFFFFF',
                )
            else:
                cell.font = Font(
                    name='Consolas',
                    size=12,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FFFFFF',
                )
        else:
            cell.font = Font(
                name='Consolas',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000',
            )
        if ',' in cell.value:
            cell.value = cell.value.replace(',', '\n')
            cell.alignment = Alignment(
                horizontal='left', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0
            )
        else:
            cell.alignment = Alignment(
                horizontal='left', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0
            )

    @staticmethod
    def fix_column_dimensions(ws):
        dims = dict()
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if '\n' in cell.value:
                        for item in cell.value.split('\n'):
                            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(item))))
                    else:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 10

    @staticmethod
    def freeze_excel_toprow(worksheet):
        worksheet.freeze_panes = 'A2'

    @staticmethod
    def get_missing_items(src: List, dst: List):
        """Get a set that represents the diff between two lists"""
        return set(src) - set(dst)

    @staticmethod
    def resource_without_id_and_metadata(resource: Dict):
        resource_without_metadata = copy.deepcopy(resource)
        resource_without_metadata.pop('links', None)
        resource_without_metadata.pop('metadata', None)
        resource_without_metadata.pop('id', None)
        return resource_without_metadata

    @staticmethod
    def diff_empty(diff: Dict):
        """Verify if a diff is empty"""
        empty = True
        for item in diff.values():
            if len(item) != 0:
                return False
        return empty

    @staticmethod
    def staticroute_count(routes: List):
        """
        Parse list of staticroute api payloads and return no. of routes across all api payloads
        """
        count = 0
        for route in routes:
            count += len(route['selectedNetworks'])
        return count

    def get_subinterface_diff(self, src: List, dst: List):
        """Generate a dictionary containing the api payloads neccessary to sync subinterface configs
        between two ftd ha pairs.
        """
        src = [item for item in src if 'readOnly' not in item['metadata']]
        dst = [item for item in dst if 'readOnly' not in item['metadata']]
        src_simplified = [item['ifname'] for item in src]
        dst_simplified = [item['ifname'] for item in dst]
        common = set(src_simplified) & set(dst_simplified)
        diff = {
            'create': [
                item for item in src if item['ifname'] in list(self.get_missing_items(src_simplified, dst_simplified))
            ],
            'update': [],
            'delete': [
                item for item in dst if item['ifname'] in list(self.get_missing_items(dst_simplified, src_simplified))
            ],
        }
        for src_intf in src:
            if src_intf['ifname'] in common:
                for dst_intf in dst:
                    if src_intf['ifname'] == dst_intf['ifname']:
                        dst_uuid = dst_intf['id']
                        if self.resource_without_id_and_metadata(src_intf) != self.resource_without_id_and_metadata(
                            dst_intf
                        ):
                            item_to_update = src_intf
                            item_to_update['id'] = dst_uuid
                            diff['update'].append(item_to_update)
        return diff

    def get_monitoredinterface_diff(self, src: List, dst: List):
        """Generate a dictionary containing the api payloads neccessary to sync monitoredinterface configs
        between two ftd ha pairs
        """
        diff = {'update': []}
        for dst_item in dst:
            for src_item in src:
                if dst_item['name'] == src_item['name']:
                    if self.resource_without_id_and_metadata(dst_item) != self.resource_without_id_and_metadata(
                        src_item
                    ):
                        updated_item = src_item
                        updated_item['id'] = dst_item['id']
                        diff['update'].append(updated_item)
        return diff

    @staticmethod
    def get_staticroute_diff_create(src: List, dst: List):
        """helper function that returns a list of staticroutes that must be created on the destination
        device to be in sync with the source device
        """
        create = list()
        for src_item in src:
            exists = False
            for dst_item in dst:
                if (
                    src_item['interfaceName'] == dst_item['interfaceName']
                    and src_item['gateway'] == dst_item['gateway']
                ):
                    exists = True
            if not exists:
                create.append(src_item)
        return create

    @staticmethod
    def get_staticroute_diff_delete(src: List, dst: List):
        """helper function that returns a list of staticroutes that must be deleted on the destination
        device to be in sync with the source device
        """
        delete = list()
        for dst_item in dst:
            exists = False
            for src_item in src:
                if (
                    src_item['interfaceName'] == dst_item['interfaceName']
                    and src_item['gateway'] == dst_item['gateway']
                ):
                    exists = True
            if not exists:
                delete.append(dst_item)
        return delete

    @staticmethod
    def get_staticroute_diff_update(src: List, dst: List):
        """helper function that returns a list of staticroutes that must be updated on the destination
        device to be in sync with the source device
        """
        update = list()
        for dst_item in dst:
            for src_item in src:
                if (
                    src_item['interfaceName'] == dst_item['interfaceName']
                    and src_item['gateway'] == dst_item['gateway']
                ):
                    if src_item['selectedNetworks'] != dst_item['selectedNetworks']:
                        updated_item = dst_item.copy()
                        updated_item['selectedNetworks'] = src_item['selectedNetworks']
                        update.append(updated_item)
        return update

    def get_staticroute_diff(self, src: List, dst: List):
        """Generate a dictionary containing the api payloads neccessary to sync staticroute configs
        between two ftd ha pairs
        """
        diff = {
            'create': self.get_staticroute_diff_create(src, dst),
            'update': self.get_staticroute_diff_update(src, dst),
            'delete': self.get_staticroute_diff_delete(src, dst),
        }
        return diff

    @staticmethod
    def get_policyassignment_diff(policyassignments: List, src_device_id: str, dst_device_id):
        """helper function to determine which policies are applied to the source device and
        therefore must be applied to the destination device
        """
        diff = {'update': list()}
        for policyassignment in policyassignments:
            src_match = False
            dst_match = False
            for target in policyassignment['targets']:
                if target['id'] == src_device_id:
                    src_match = True
                if target['id'] == dst_device_id:
                    dst_match = True
            if src_match and not dst_match:
                updated_item = policyassignment
                updated_item['targets'].append({'id': dst_device_id, 'type': 'DeviceHAPair'})
                diff['update'].append(updated_item)
        return diff

    def sync_device_subinterfaces(self, src_device_id: str, dst_device_id: str, dry_run=False):
        """Synchronize the subinterface configuration between two ftd ha pairs
        """
        src_interfaces = self.fmc.device.devicerecord.subinterface.get(container_uuid=src_device_id)
        dst_interfaces = self.fmc.device.devicerecord.subinterface.get(container_uuid=dst_device_id)
        diff = self.get_subinterface_diff(src_interfaces, dst_interfaces)
        if not self.diff_empty(diff):
            logger.info(
                (
                    'Subinterface configuration is out of sync. ',
                    '%s interfaces will be created, %s will be updated and %s will be deleted',
                ),
                len(diff['create']),
                len(diff['update']),
                len(diff['delete']),
            )
            if dry_run:
                logger.info('The following changes would be performed:\n%s', json.dumps(diff, indent=4, sort_keys=True))
                return
            for item in diff['create']:
                logger.info('Trying to create interface %s (%s.%s)', item['ifname'], item['name'], item['subIntfId'])
                self.fmc.device.devicerecord.subinterface.create(container_uuid=dst_device_id, data=item)
                logger.info(
                    'Successfully created interface %s (%s.%s)', item['ifname'], item['name'], item['subIntfId']
                )

            for item in diff['update']:
                logger.info('Trying to update interface %s (%s.%s)', item['ifname'], item['name'], item['subIntfId'])
                self.fmc.device.devicerecord.subinterface.update(container_uuid=dst_device_id, data=item)
                logger.info(
                    'Successfully updated interface %s (%s.%s)', item['ifname'], item['name'], item['subIntfId']
                )

            for item in diff['delete']:
                try:
                    logger.info(
                        'Trying to delete interface %s (%s.%s)', item['ifname'], item['name'], item['subIntfId']
                    )
                    self.fmc.device.devicerecord.subinterface.delete(container_uuid=dst_device_id, uuid=item['id'])
                    logger.info(
                        'Successfully deleted interface %s (%s.%s)', item['ifname'], item['name'], item['subIntfId']
                    )
                except GenericApiError as exc:
                    logger.info(
                        'Failed to delete interface %s (%s.%s)', item['ifname'], item['name'], item['subIntfId']
                    )
                    logger.error(str(exc))

        else:
            logger.info('Subinterface configuration is in sync. Skipping synchronization')

    def sync_device_monitoredinterfaces(self, src_device_id: str, dst_device_id: str, dry_run=False):
        """Synchronize the monitoredinterfaces configuration between two ftd ha pairs
        """
        src_interfaces = self.fmc.devicehapair.ftdhapair.monitoredinterface.get(container_uuid=src_device_id)
        dst_interfaces = self.fmc.devicehapair.ftdhapair.monitoredinterface.get(container_uuid=dst_device_id)
        diff = self.get_monitoredinterface_diff(src_interfaces, dst_interfaces)
        if not self.diff_empty(diff):
            logger.info(
                'Monitored interface configuration is out of sync. %s interfaces will be updated', len(diff['update'])
            )
            if dry_run:
                logger.info('The following changes would be performed:\n', json.dumps(diff, indent=4, sort_keys=True))
                return
            for item in diff['update']:
                logger.info('Trying to update interface %s', item['name'])
                self.fmc.devicehapair.ftdhapair.monitoredinterface.update(container_uuid=dst_device_id, data=item)
                logger.info('Successfully updated interface %s', item['name'])
        else:
            logger.info('Monitored interface configuration is in sync. Skipping synchronization')

    def sync_device_staticroutes(self, src_device_id: str, dst_device_id: str, dry_run=False):
        """Synchronize the static routing configuration between two ftd ha pairs
        """
        src = self.fmc.device.devicerecord.routing.ipv4staticroute.get(container_uuid=src_device_id)
        dst = self.fmc.device.devicerecord.routing.ipv4staticroute.get(container_uuid=dst_device_id)
        diff = self.get_staticroute_diff(src, dst)
        if not self.diff_empty(diff):
            logger.info(
                'Configuration is out of sync.%s ipv4 staticroutes will be created, '
                '%s  will be updated and %s will be deleted',
                self.staticroute_count(diff['create']),
                self.staticroute_count(diff['update']),
                self.staticroute_count(diff['delete']),
            )
            if dry_run:
                logger.info('The following changes would be performed:\n%s', json.dumps(diff, indent=4, sort_keys=True))
                return
            for item in diff['create']:
                gateway = (
                    item['gateway']['object']['name']
                    if 'object' in item['gateway']
                    else item['gateway']['literal']['value']
                )
                interface = item['interfaceName']
                networks = ', '.join([obj['name'] for obj in item['selectedNetworks']])
                logger.info('Trying to create IPv4 static route to %s via %s (%s)', networks, gateway, interface)
                self.fmc.device.devicerecord.routing.ipv4staticroute.create(container_uuid=dst_device_id, data=item)
                logger.info('Successfully created IPv4 static route')
            for item in diff['update']:
                gateway = (
                    item['gateway']['object']['name']
                    if 'object' in item['gateway']
                    else item['gateway']['literal']['value']
                )
                interface = item['interfaceName']
                networks = ', '.join([obj['name'] for obj in item['selectedNetworks']])
                logger.info('Trying to update IPv4 static route to %s via %s (%s)', networks, gateway, interface)
                self.fmc.device.devicerecord.routing.ipv4staticroute.update(container_uuid=dst_device_id, data=item)
                logger.info('Successfully updated IPv4 static route to %s via %s (%s)', networks, gateway, interface)
            for item in diff['delete']:
                gateway = (
                    item['gateway']['object']['name']
                    if 'object' in item['gateway']
                    else item['gateway']['literal']['value']
                )
                interface = item['interfaceName']
                networks = ', '.join([obj['name'] for obj in item['selectedNetworks']])
                logger.info('Trying to delete IPv4 static route to %s via %s (%s)', networks, gateway, interface)
                self.fmc.device.devicerecord.routing.ipv4staticroute.delete(
                    container_uuid=dst_device_id, uuid=item['id']
                )
                logger.info('Successfully deleted IPv4 static route')
        else:
            logger.info('IPv4 static routing configuration is in sync. Skipping synchronization')

    def sync_device_policyassignments(self, src_device_id: str, dst_device_id: str, dry_run=False):
        """Synchronize the policy assignments between two ftd ha pairs
        """
        policyassignments = self.fmc.assignment.policyassignment.get()
        diff = self.get_policyassignment_diff(policyassignments, src_device_id, dst_device_id)
        if not self.diff_empty(diff):
            logger.info(
                'Policy assignment configuration is out of sync. %s policy assignments will be updated',
                len(diff['update']),
            )
            if dry_run:
                logger.info('The following changes would be performed:\n%s', json.dumps(diff, indent=4, sort_keys=True))
                return
            for item in diff['update']:
                policy_type = item['policy']['type']
                policy_name = item['policy']['name']
                logger.info('Trying to update policy assignment for %s %s', policy_type, policy_name)
                self.fmc.assignment.policyassignment.update(data=item)
                logger.info('Successfully updated policy assignment')
        else:
            logger.info('Policy assignments are in sync. Skipping synchronization')

    def deploy_configuration(self, device_id: str):
        """Deploy configuration from firepower management center to device
        """
        deployment_pending = False
        deployments_pending = self.fmc.deployment.deployabledevices.get()
        for device in deployments_pending:
            if device_id == device['device']['id']:
                device_name = device['name']
                deployment_pending = True
        if deployment_pending:
            if self.fmc.conn.dry_run:
                logger.info('Configuration deployment pending for %s. Skipping due to dry-run', device_name)
                return
            logger.info('Configuration deployment pending for %s. Initiating deployment process', device_name)
            data = {
                'forceDeploy': True,
                'ignoreWarning': True,
                'deviceList': [device_id],
                'type': 'DeploymentRequest',
                'version': int(1000 * datetime.now().timestamp()),
            }
            self.fmc.deployment.deploymentrequest.create(data=data)
            logger.info('Successfully scheduled configuration deployment')

    def expanded_accessrules(self, accessrules, objects, device):
        accessrules = [benedict(accessrule) for accessrule in accessrules]
        for accessrule in accessrules:
            if 'sourceNetworks.objects' in accessrule:
                for k, v in enumerate(accessrule['sourceNetworks']['objects']):
                    accessrule['sourceNetworks']['objects'][k] = self.expanded_obj(v, objects, device)
            if 'destinationNetworks.objects' in accessrule:
                for k, v in enumerate(accessrule['destinationNetworks']['objects']):
                    accessrule['destinationNetworks']['objects'][k] = self.expanded_obj(v, objects, device)
            if 'sourcePorts.objects' in accessrule:
                for k, v in enumerate(accessrule['sourcePorts']['objects']):
                    accessrule['sourcePorts']['objects'][k] = self.expanded_obj(v, objects, device)
            if 'destinationPorts.objects' in accessrule:
                for k, v in enumerate(accessrule['destinationPorts']['objects']):
                    accessrule['destinationPorts']['objects'][k] = self.expanded_obj(v, objects, device)
            if 'urls.objects' in accessrule:
                for k, v in enumerate(accessrule['urls']['objects']):
                    accessrule['urls']['objects'][k] = self.expanded_obj(v, objects, device)
        return accessrules

    def expanded_obj(self, obj: Dict, objects: Dict, device: str):
        for item in objects[obj['type'].lower()]:
            if item['id'] == obj['id']:
                if 'group' not in obj['type'].lower():
                    return self.override_obj(item, device)
                else:
                    obj = self.override_obj(item, device)
                    if 'objects' in obj:
                        for nested_obj_id, nested_obj in enumerate(obj['objects']):
                            obj['objects'][nested_obj_id] = self.expanded_obj(nested_obj, objects, device)
        return obj

    @staticmethod
    def override_obj(obj: Dict, device: str):
        if 'overrides' in obj:
            for override in obj['overrides']:
                if override['overrides']['target']['id'] == device:
                    if 'group' in obj['type']:
                        obj['literals'] = list()
                        obj['objects'] = list()
                        if 'literals' in override:
                            obj['literals'] = override['literals']
                        if 'objects' in override:
                            obj['objects'] = override['objects']
                    else:
                        obj['value'] = override['value']
        return obj

    def match_protocol(self, accessrule: Dict, protocol: str):
        protocol_map = {'http': [80, 443], 'smtp': [25], 'smb': [445], 'ftp': [20, 21]}

        # no need to check protocols if protocol is set to "any"
        if protocol == 'any':
            return True
        # no need to check further if there are no destination ports specified in the rule
        if 'destinationPorts' not in accessrule:
            return False
        # copy destinationPorts and search for matches in objects
        dst_ports = copy.deepcopy(accessrule['destinationPorts'])
        if 'objects' in dst_ports:
            for obj in dst_ports['objects']:
                if obj['type'] == 'PortObjectGroup':
                    parent_obj = self.fmc.object.portobjectgroup.get(uuid=obj['id'])
                    for child_obj in parent_obj['objects']:
                        expanded_child_obj = self.fmc.object.protocolportobject.get(uuid=child_obj['id'])
                        if expanded_child_obj['protocol'] == 'TCP':
                            if int(expanded_child_obj['port']) in protocol_map[protocol]:
                                return True
                else:
                    expanded_obj = self.fmc.object.protocolportobject(uuid=obj['id'])
                    if expanded_obj['protocol'] == 'TCP':
                        if int(expanded_obj['port']) in protocol_map[protocol]:
                            return True
        # check for match in literals
        if 'literals' in dst_ports:
            for literal in dst_ports['literals']:
                if literal['protocol'] == '6':
                    if int(literal['port']) in protocol_map[protocol]:
                        return True
        return False

    def accesspolicy_filepolicy_enable(self, accesspolicy, filepolicy, protocol):
        accessrules = self.fmc.policy.accesspolicy.accessrule.get(container_uuid=accesspolicy)
        accessrules = [item for item in copy.deepcopy(accessrules) if self.match_protocol(item, protocol)]
        for accessrule in accessrules:
            if 'filePolicy' in accessrule:
                if accessrule['filePolicy']['id'] == filepolicy:
                    logger.info('Filepolicy is already applied to accessrule "%s". Skipping.', accessrule['name'])
                else:
                    logger.info(
                        'Accessrule "%s" matches protocol "%s". Applying filepolicy to rule',
                        accessrule['name'],
                        protocol,
                    )
                    accessrule['filePolicy'] = {'id': filepolicy}
                    self.fmc.policy.accesspolicy.accessrule.update(container_uuid=accesspolicy, data=accessrule)
            else:
                logger.info(
                    'Accessrule "%s" matches protocol "%s". Applying filepolicy to rule', accessrule['name'], protocol
                )
                accessrule['filePolicy'] = {'id': filepolicy}
                self.fmc.policy.accesspolicy.accessrule.update(container_uuid=accesspolicy, data=accessrule)
